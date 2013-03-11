'''
Created on Mar 9, 2013

@author: me
'''
import csv
from openpyxl import load_workbook

class Xlsx:
    '''
    classdocs
    '''


    def __init__(self, xlsxfile):
        '''
        Constructor
        '''
        self.xlsxfile = xlsxfile
        
    def export_sheet_to_csv(self, sheetname, csvfile):
        '''
        Open Test Cases file
        '''

        with open(csvfile, 'wb') as f:
            writer = csv.writer(f)

            workbook = load_workbook(filename = self.xlsxfile, use_iterators = True)
            worksheet = workbook.get_sheet_by_name(name = sheetname)
            
            
            for row in worksheet.iter_rows():
                row_list = []
                for cell in row:
                    try:
                        cell_value = cell.internal_value
                        cell_value = int(cell_value)
                    except:
                        pass
                    row_list.append(cell_value)
                writer.writerow(row_list)

if __name__ == "__main__":
    
    pass